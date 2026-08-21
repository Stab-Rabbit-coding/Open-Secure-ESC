#!/usr/bin/env python3
"""Add the Motor and Shaft Sensor axis sheets to docs/decision-matrix.xlsx.

Governed by AGENTS.md. The workbook shipped with five axis sheets (Voltage,
Amperage, Protocol, Control, EMI Hardening), but a build cannot actually be
instantiated from it without two more:

  * **Motor** -- brushed vs brushless. This is NOT a parameter tweak. It
    changes the bridge topology, the switch-position count, the gate-driver
    PART NUMBER and the shunt count. A 3-phase gate driver cannot drive an
    H-bridge as-is.
  * **Shaft Sensor** -- sensorless / Hall / quadrature encoder / resolver.
    Drives the MCU pin budget, the connector, and whether an extra IC
    (a resolver-to-digital converter) joins the BOM at all.

See docs/solutions/architecture-patterns/esc-build-instantiation-workflow.md
for why these two are the axes that were missing.

WHAT THIS SCRIPT WILL NOT DO
-----------------------------
It does not invent part numbers. AGENTS.md Sec.1.3 forbids filling a cell with
a plausible-sounding part, and several cells below genuinely have no verified
answer in this repo yet:

  * No H-bridge gate driver has been sourced or verified here. The brushed
    row's driver cell is `Open / unresolved`, not a guess.
  * No resolver-to-digital converter has been sourced or verified.
  * No Hall-sensor part has been verified. The Control sheet already names
    TI DRV5013 as a *fallback candidate* for its sensorless rows, so that
    name is carried across as a candidate, with its existing status, rather
    than promoted.

Every such cell reads `TBD -- requires primary-source verification
(AGENTS.md Sec.1.3)` and carries status `Open / unresolved`. That is the
correct state for this workbook, not a defect in it.

The script is idempotent: re-running replaces the two sheets rather than
duplicating them, and it backs the workbook up first.

Requires: pip install openpyxl

Usage:
    python3 docs/tools/add_motor_and_shaft_sensor_sheets.py [--dry-run]
"""
# Authored by Claude Opus 5 (Anthropic) 2026-08-16. AI-generated; part
# selections are carried from this repo's own verified records or explicitly
# left open. Not human-authored.

import argparse
import shutil
from datetime import datetime, timezone
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

REPO = Path(__file__).resolve().parents[2]
WORKBOOK = REPO / "docs" / "decision-matrix.xlsx"

HEADER_FILL = PatternFill("solid", fgColor="FF2F5597")
HEADER_FONT = Font(bold=True, size=10, color="FFFFFFFF")
TITLE_FONT = Font(bold=True, size=16)
SUB_FONT = Font(size=10)
BODY_ALIGN = Alignment(wrap_text=True, vertical="top")
HEADER_ALIGN = Alignment(wrap_text=True, vertical="center")

TBD = "TBD -- requires primary-source verification (AGENTS.md Sec.1.3)"

# ---------------------------------------------------------------- Motor -----
MOTOR = {
    "title": "Motor Type Decision Matrix",
    "subtitle": "Build axis: Motor (Brushed / Brushless)",
    "widths": [18, 26, 14, 34, 24, 14, 34, 18, 22],
    "columns": [
        "Motor Type",
        "Bridge Topology",
        "Switch Positions",
        "Gate Driver Part",
        "Output Terminals",
        "Shunt Qty",
        "Commutation Requirement",
        "REFERENCES.md Tags",
        "Status",
    ],
    "rows": [
        [
            "Brushless (BLDC / PMSM)",
            "Three half-bridges (3-phase inverter)",
            "6 (3 high + 3 low)",
            "TI DRV8353S, 3-phase smart gate driver, SPI variant [21]",
            "3 (PH_A / PH_B / PH_C)",
            "3 (one per phase, low-side)",
            "Six-step trapezoidal or FOC; rotor position from the Shaft Sensor axis",
            "[21]",
            "Verified (local PDF)",
        ],
        [
            "Brushed (DC)",
            "Single H-bridge",
            "4 (2 high + 2 low)",
            TBD + ". No H-bridge gate driver has been sourced or verified in "
            "this repo. NOTE, not yet vetted: a 3-phase driver such as the "
            "DRV8353S [21] can drive an H-bridge by using two of its three "
            "half-bridges and leaving the third unused -- this reuses an "
            "already-verified part at the cost of wasted silicon and board "
            "area, and must be checked against [21] before adoption.",
            "2 (M+ / M-)",
            "1-2 (bridge-leg or in-line)",
            (
                "None -- direction and magnitude set by bridge polarity and duty "
                "cycle; no rotor position needed for basic operation"
            ),
            "-",
            "Open / unresolved",
        ],
    ],
}

# --------------------------------------------------------- Shaft Sensor -----
SHAFT = {
    "title": "Shaft Sensor Decision Matrix",
    "subtitle": "Build axis: Shaft Sensor (None-sensorless / Hall / "
    "Quadrature encoder / Resolver)",
    "widths": [22, 30, 22, 34, 22, 38, 18, 22],
    "columns": [
        "Sensor Type",
        "Rotor Position Source",
        "MCU Pins Required",
        "Additional BOM Component",
        "Connector",
        "Firmware Workflow Steps",
        "REFERENCES.md Tags",
        "Status",
    ],
    "rows": [
        [
            "None (sensorless)",
            (
                "Back-EMF / observer computed from phase voltages and currents; "
                "same observer the Control sheet's closed-loop rows already use "
                "(TI SLAAE96A [13])"
            ),
            (
                "3 ADC (one per phase voltage divider). Reuses the current-sense "
                "ADC channels already allocated on the Amperage axis."
            ),
            (
                "3 resistor dividers (6 R) plus filter capacitors, one per phase. "
                "Generic passives, values sized against the pack voltage of the "
                "chosen Voltage tier."
            ),
            "None -- no sensor cable leaves the board",
            (
                "Implement the observer per [13]; handle open-loop startup ramp, "
                "since back-EMF is unusable near zero speed."
            ),
            "[13]",
            "Candidate (unverified)",
        ],
        [
            "Hall effect (3-sensor, 120 deg)",
            (
                "Three digital Hall outputs giving six commutation states per "
                "electrical revolution"
            ),
            "3 digital in (interrupt-capable), plus a sensor supply rail",
            (
                "3 pull-up resistors plus ESD protection on the cable entry. "
                "Sensor part: TI DRV5013 is carried from the Control sheet as a "
                "fallback candidate, at that sheet's existing status -- it has "
                "not been separately verified for this axis."
            ),
            "5-pin (VCC, GND, HA, HB, HC)",
            (
                "Decode the 3-bit Hall state into a 6-step commutation table; "
                "derive speed from state-transition timing."
            ),
            "[1]",
            "Open / unresolved",
        ],
        [
            "Quadrature encoder (incremental)",
            "Two phase-quadrature channels A/B, plus optional once-per-rev index Z",
            (
                "2 digital in, or 3 with index. Prefer a timer with a hardware "
                "quadrature decoder if the chosen MCU has one."
            ),
            (
                "Pull-ups for an open-collector encoder, or a differential line "
                "receiver for RS-422 encoders -- " + TBD
            ),
            "4-pin (VCC, GND, A, B) or 6-pin with Z and shield",
            (
                "Configure the hardware quadrature decoder; establish a zero "
                "reference from the index pulse or a homing routine."
            ),
            "-",
            "Open / unresolved",
        ],
        [
            "Resolver",
            (
                "Ratiometric SIN/COS windings excited by an AC reference; "
                "absolute position within one electrical revolution"
            ),
            (
                "Depends on the converter: an SPI or parallel bus to a "
                "resolver-to-digital converter, plus an excitation output"
            ),
            (
                "Resolver-to-digital converter IC and an excitation amplifier -- "
                + TBD
                + ". This is the only shaft-sensor option that adds a whole "
                "IC to the BOM."
            ),
            "6-pin (EXC+, EXC-, SIN+, SIN-, COS+, COS-), shielded",
            (
                "Drive the excitation reference; read absolute angle from the "
                "converter; no startup ramp needed, unlike sensorless."
            ),
            "-",
            "Open / unresolved",
        ],
    ],
}


def write_sheet(wb, name: str, spec: dict) -> None:
    """Create (or replace) one axis sheet in the shipped house style."""
    if name in wb.sheetnames:
        del wb[name]
    ws = wb.create_sheet(name)

    ws["A1"] = spec["title"]
    ws["A1"].font = TITLE_FONT
    ws["A1"].alignment = Alignment(vertical="center")
    ws["A2"] = spec["subtitle"]
    ws["A2"].font = SUB_FONT

    for col, header in enumerate(spec["columns"], start=1):
        c = ws.cell(row=4, column=col, value=header)
        c.fill, c.font, c.alignment = HEADER_FILL, HEADER_FONT, HEADER_ALIGN

    for r, row in enumerate(spec["rows"], start=5):
        for col, value in enumerate(row, start=1):
            c = ws.cell(row=r, column=col, value=value)
            c.alignment = BODY_ALIGN

    for col, width in enumerate(spec["widths"], start=1):
        ws.column_dimensions[ws.cell(row=4, column=col).column_letter].width = width
    ws.freeze_panes = "A5"


def update_legend(wb) -> None:
    """Add the two new sheets to the Legend's sheet list."""
    ws = wb["Legend"]
    existing = {
        str(ws.cell(row=r, column=1).value or "").strip()
        for r in range(1, ws.max_row + 1)
    }
    row = ws.max_row + 1
    additions = [
        (
            "Motor",
            (
                "Brushed vs brushless: bridge topology, switch-position "
                "count, gate-driver part, shunt count. Changes the power "
                "stage, not just a setting."
            ),
        ),
        (
            "Shaft Sensor",
            (
                "Sensorless / Hall / quadrature encoder / resolver: "
                "MCU pin budget, connector, and any added IC."
            ),
        ),
    ]
    for label, desc in additions:
        if label in existing:
            continue
        ws.cell(row=row, column=1, value=label).font = Font(size=10)
        ws.cell(row=row, column=2, value=desc).alignment = BODY_ALIGN
        row += 1
    note_row = row + 1
    ws.cell(
        row=note_row,
        column=1,
        value=(
            f"Amended {datetime.now(timezone.utc).date().isoformat()}: added the Motor "
            "and Shaft Sensor axis sheets. Cells reading 'TBD -- requires "
            "primary-source verification' are genuinely unsourced, not "
            "placeholders to be filled in from memory (AGENTS.md Sec.1.3)."
        ),
    ).font = Font(size=10, italic=True)


def main() -> int:
    """Add both sheets to the workbook."""
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--dry-run", action="store_true")
    args = ap.parse_args()

    wb = openpyxl.load_workbook(WORKBOOK)
    before = list(wb.sheetnames)

    write_sheet(wb, "Motor", MOTOR)
    write_sheet(wb, "Shaft Sensor", SHAFT)
    update_legend(wb)

    # Keep axis sheets in build-walk order, Legend first.
    order = [
        "Legend",
        "Voltage",
        "Amperage",
        "Motor",
        "Shaft Sensor",
        "Protocol",
        "Control",
        "EMI Hardening",
    ]
    wb._sheets = [wb[n] for n in order if n in wb.sheetnames] + [
        wb[n] for n in wb.sheetnames if n not in order
    ]

    print(f"sheets before: {before}")
    print(f"sheets after:  {wb.sheetnames}")
    if args.dry_run:
        print("dry run -- nothing written")
        return 0

    backup = WORKBOOK.with_suffix(".xlsx.bak")
    shutil.copy2(WORKBOOK, backup)
    wb.save(WORKBOOK)
    print(f"backed up  {backup.name}")
    print(f"wrote      {WORKBOOK}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
