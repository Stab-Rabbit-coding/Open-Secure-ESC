#!/usr/bin/env python3
"""Add the Form Factor axis sheet to docs/decision-matrix.xlsx.

Governed by AGENTS.md. Companion to docs/tools/strip_width.py.

WHY THIS AXIS EXISTS
--------------------
Every build so far has assumed a flat rigid board in free space. Mounting one
inside a curved host -- a nacelle, a boom, a tube -- makes the board a CHORD
of the bore, and the gap between chord and arc (the sagitta) is radial depth
stolen from whatever else shares that annulus. Get it wrong and the skin does
not close, so the number belongs in the decision matrix as an axis value, not
in someone's head during placement.

THE THREE VALUES THIS SHEET CARRIES, AND WHY
---------------------------------------------
Max board width and max board length alone do not describe a curved fit,
because a faceted board's governing dimension is the width of ONE rigid facet
-- the strip width -- which the host radius and depth budget fix:

    w = 2 * sqrt(2*R*s - s^2)        widest strip fitting radial depth s
    s = R - sqrt(R^2 - (w/2)^2)      depth consumed by a strip of width w

So the sheet carries strip width and sagitta as a pair (each derives the
other) alongside the overall envelope.

WIDTH AND LENGTH ARE COUPLED, AND THE SHEET SAYS SO
----------------------------------------------------
A board holds a fixed set of parts. Pinning one dimension drives the other,
and narrowing usually costs more length than the width it saves. The
`Component-Driven Dimension` column records which dimension the enclosure
pins and which the parts then dictate. On this design, at 24.00 mm:

  * isolation_envelope.py: below its 31.86 mm minimum the control section can
    no longer sit BETWEEN the isolated rows and must stack beyond them in Y,
    priced at +19.50 mm of length.
  * the FET bridge spans 26.96 mm across Q1-Q6 in its 3x2 arrangement, which
    does not fit a 24.00 mm strip at all; 2x3 adds further length.

WHAT THIS SCRIPT WILL NOT DO
-----------------------------
It does not invent flex specifications. **No flex or rigid-flex standard is
catalogued in REFERENCES.md** -- neither IPC-2223 (flex design) nor IPC-6013
(flex qualification) has been obtained -- so bend radius, flex-zone layer
count and copper limits are recorded as TBD per AGENTS.md Sec.1.3 rather than
filled with plausible numbers.

The script is idempotent: re-running replaces the sheet rather than
duplicating it, and it backs the workbook up to a timestamped file.

Requires: openpyxl

Usage:
    python3 docs/tools/add_form_factor_sheet.py [--dry-run]
    python3 docs/tools/decision_matrix_to_json.py
"""
# Authored by Claude Opus 5 (Anthropic) 2026-08-31. AI-generated. The
# geometry is derivable by inspection; every part- and standard-dependent
# cell is either carried from this repo's verified records or left open.
# Not human-authored. Not yet human-reviewed.

import argparse
import math
import shutil
from datetime import datetime, timezone
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

REPO = Path(__file__).resolve().parents[2]
WORKBOOK = REPO / "docs" / "decision-matrix.xlsx"

HEADER_FILL = PatternFill("solid", fgColor="FF2F5597")
HEADER_FONT = Font(bold=True, size=10, color="FFFFFFFF")
TITLE_FONT = Font(bold=True, size=16)
SUB_FONT = Font(size=10)
BODY_ALIGN = Alignment(wrap_text=True, vertical="top")
HEADER_ALIGN = Alignment(wrap_text=True, vertical="center")

TBD = "TBD -- requires primary-source verification (AGENTS.md Sec.1.3)"

# Reference host geometry: the 30 mm-radius nacelle bore this axis was raised
# for. Values are recomputed here rather than transcribed, so the sheet cannot
# drift from docs/tools/strip_width.py.
REF_RADIUS_MM = 30.0
REF_DEPTH_MM = 2.5
REF_WIDTH_MM = 2.0 * math.sqrt(2.0 * REF_RADIUS_MM * REF_DEPTH_MM - REF_DEPTH_MM**2)
NOMINAL_WIDTH_MM = 24.0
NOMINAL_SAGITTA_MM = REF_RADIUS_MM - math.sqrt(
    REF_RADIUS_MM**2 - (NOMINAL_WIDTH_MM / 2.0) ** 2
)
NOMINAL_FACET_DEG = 2.0 * math.degrees(
    math.asin((NOMINAL_WIDTH_MM / 2.0) / REF_RADIUS_MM)
)

PREAMBLE = [
    ("Width from radial depth", "w = 2 * sqrt(2*R*s - s^2)"),
    ("Radial depth from width", "s = R - sqrt(R^2 - (w/2)^2)"),
    ("Reference host bore radius R (mm)", REF_RADIUS_MM),
    ("Reference radial depth budget s (mm)", REF_DEPTH_MM),
    ("Resulting max strip width w (mm)", round(REF_WIDTH_MM, 2)),
]

FORM_FACTOR = {
    "title": "Form Factor Decision Matrix",
    "subtitle": "Build axis: Form Factor (Flat / Faceted rigid-flex / Faceted separate boards). "
    "Strip width and sagitta derive from each other; see the preamble.",
    "widths": [22, 32, 26, 26, 32, 26, 42, 34, 46, 16, 20],
    "columns": [
        "Form Factor",
        "Host Geometry",
        "Max Strip Width (mm)",
        "Sagitta / Radial Depth (mm)",
        "Max Board Width (mm)",
        "Max Board Length (mm)",
        "Component-Driven Dimension",
        "Facet Interconnect",
        "Workflow Steps",
        "REFERENCES.md Tags",
        "Status",
    ],
    "rows": [
        [
            "Flat (default)",
            "Planar mounting, no host curvature. Sagitta does not apply.",
            "n/a -- one rigid panel",
            "n/a",
            (
                "No host limit. The LOWER bound of 31.86 mm is set by [9] "
                "Table 6 creepage via docs/tools/isolation_envelope.py, not "
                "by any enclosure."
            ),
            "Driven by components.",
            (
                "Length. Width is pinned at the creepage minimum, so the part "
                "count sets the length."
            ),
            "None -- a single rigid board.",
            (
                "Default; no extra step. Both existing builds are this form "
                "factor: builds/6s/50A/CAN_485_faraday/ at 32.00 x 66.10 mm "
                "and .../CAN_485_faraday_sameend/ at 32.00 x 76.10 mm."
            ),
            "[9]",
            "Candidate (unverified)",
        ],
        [
            "Faceted rigid-flex",
            (
                f"Cylindrical bore of radius R. Each rigid facet is a flat "
                f"chord; its sagitta is the radial depth it consumes from the "
                f"annulus. At the reference R = {REF_RADIUS_MM:.1f} mm a "
                f"{NOMINAL_WIDTH_MM:.2f} mm facet subtends "
                f"{NOMINAL_FACET_DEG:.2f} deg, so 4 facets cover 180 deg."
            ),
            (
                f"{NOMINAL_WIDTH_MM:.2f} mm nominal at the reference geometry "
                f"({REF_WIDTH_MM:.2f} mm exact for s = {REF_DEPTH_MM:.1f} mm, "
                f"rounded down). Recompute per host: "
                f"docs/tools/strip_width.py --radius R --depth s"
            ),
            (
                f"{NOMINAL_SAGITTA_MM:.3f} mm at w = {NOMINAL_WIDTH_MM:.2f} "
                f"mm, R = {REF_RADIUS_MM:.1f} mm. This is the depth the board "
                f"steals from the annulus -- budget it against the bore "
                f"contents, not against the skin alone."
            ),
            "Equals the strip width for a single-facet-wide board.",
            (
                "Driven by components; the sum of the facet lengths plus the "
                "flex hinges between them."
            ),
            (
                "Length, and it grows faster than the width shrinks. At "
                "24.00 mm on THIS design: isolation_envelope.py can no longer "
                "seat the control section between the isolated rows and "
                "stacks it in Y instead, +19.50 mm; and the FET bridge spans "
                "26.96 mm across Q1-Q6 as a 3x2, so it does not fit the strip "
                "at all and must become 2x3, adding more. Estimated "
                "24 x 105-115 mm -- a new layout sharing the BOM, NOT a "
                "variant of the flat board."
            ),
            (
                "Flex hinge -- "
                + TBD
                + ". No flex or rigid-flex standard is catalogued in "
                "REFERENCES.md; neither IPC-2223 (flex design) nor IPC-6013 "
                "(flex qualification) has been obtained. Bend radius, "
                "flex-zone layer count and copper limits cannot be stated "
                "until one is."
            ),
            (
                "Acquire and cite a flex design standard FIRST (AGENTS.md "
                "Sec.2). Then: run strip_width.py for the actual host radius "
                "and depth budget; re-run isolation_envelope.py at the target "
                "width to price the length; re-lay the FET bridge to fit the "
                "strip; keep the whole 50 A power stage on ONE rigid facet -- "
                "flex copper is thin and polyimide is a poor thermal path, "
                "and conductor_sizing.py already requires 2 oz just for the "
                "phase pours; re-run conductor_sizing.py for the new "
                "geometry."
            ),
            "[9]",
            "Open / unresolved",
        ],
        [
            "Faceted separate boards",
            (
                "Same chord geometry as rigid-flex -- each rigid panel is a "
                "flat chord of the bore -- but the panels are independent "
                "boards joined by an interconnect rather than by a flex "
                "hinge. NOTE: the panels meet AT the bore circle, the point "
                "of maximum radius, so a connector at the joint has no "
                "outward room and protrudes inward into the bore. Budget its "
                "height against the bore contents (TODO.md 16.3)."
            ),
            (
                f"Identical to rigid-flex: {NOMINAL_WIDTH_MM:.2f} mm nominal "
                f"at the reference geometry. docs/tools/hinge_placement.py "
                f"chooses the CUT position exactly as it chooses a fold."
            ),
            (
                f"{NOMINAL_SAGITTA_MM:.3f} mm per panel, as rigid-flex, PLUS "
                f"the interconnect height at the joint vertex."
            ),
            "Equals the strip width per panel.",
            "Driven by components, as rigid-flex.",
            (
                "Length -- AND the form factor should drive the functional "
                "partition. Faceting makes logic-vs-power placement a "
                "first-class layout decision: if the whole 50 A power stage "
                "sits on ONE panel the joint carries signals only, and either "
                "form factor works. If the cut splits the power stage the "
                "joint carries up to 50 A, which only separate boards can "
                "realistically do. On the current layout the power stage "
                "spans 26.96 mm in X and logic spans 28.74 mm, overlapping "
                "over 26.96 mm, so EVERY cut splits both -- the partition "
                "must be re-oriented along the cut axis."
            ),
            (
                "Board-to-board connector, busbar, solder tab or wire link -- "
                + TBD
                + ". No interconnect part has been selected. The decisive "
                "advantage over a flex hinge is that this joint CAN carry "
                "50 A; flex copper is thin and conductor_sizing.py already "
                "requires 2 oz for the phase pours alone. The decisive "
                "disadvantage is a rigid joint in a vibration environment, "
                "where a flex hinge is compliant -- an engineering judgement "
                "here, no standard is cited for it."
            ),
            (
                "NOT blocked by the flex-standard gate (16.2): standard rigid "
                "FR-4 on both panels, any fab, full copper weight throughout. "
                "Run hinge_placement.py --optimize for the cut, then: keep "
                "the whole power stage on one panel if the interconnect is to "
                "stay signal-only; keep any rigid shielding can wholly on one "
                "panel (SH1 is 22.75 mm and nearly fills a 23.98 mm facet by "
                "itself); and either give the isolated section its own panel "
                "or let the isolation barrier fall ON the cut, where the "
                "inter-panel gap supplies the separation instead of 7.5 mm of "
                "board surface."
            ),
            "[9], [19]",
            "Open / unresolved",
        ],
    ],
}

NOTES = (
    "Choose the dimension the enclosure actually constrains and let the "
    "components drive the other -- they are not independent. A board carries "
    "a fixed part list, so narrowing it does not shrink it, it lengthens it, "
    "usually by more than the width saved. Both figures in the "
    "'Component-Driven Dimension' cells come from this repo's own tools "
    "(docs/tools/isolation_envelope.py and measurements of the Q1-Q6 "
    "placement), not from estimation. -- Curvature is NOT what blocks a "
    "narrow board on this design: creepage is. [9] Table 6 forces a 31.86 mm "
    "minimum width while the isolated transceivers sit either side of the "
    "control section, and no amount of bending changes that; only restacking "
    "the isolation section does. -- True curved rigid FR-4 does not exist: "
    "the laminate is pressed flat. 'Curved' means faceted, whether the facets "
    "are joined by flex hinges or are separate boards with an "
    "interconnect. -- CHOOSING BETWEEN THE TWO FACETED ROWS: run "
    "docs/tools/hinge_placement.py --optimize and read what crosses the cut. "
    "A signal-only joint favours rigid-flex (no connector, self-locating, "
    "compliant under vibration); a joint carrying Power-netclass conductors "
    "favours separate boards, because a flex hinge cannot realistically pass "
    "50 A. Separate boards also avoid the flex-standard acquisition gate "
    "(16.2) entirely. The tool also checks the two constraints faceting "
    "imposes on placement: a rigid shielding can may not straddle a joint, "
    "and a panel holding both sides of an isolation barrier must still meet "
    "the 31.86 mm width isolation_envelope.py derives from [9] Table 6."
)


def write_sheet(wb, name: str, spec: dict) -> None:
    """Create (or replace) the axis sheet, with a preamble block.

    Layout follows the Voltage sheet rather than the plain house style: rows
    4-8 carry key/value assumptions that decision_matrix_to_json.py picks up
    via read_preamble(), and the header row therefore sits at row 10. The
    exporter locates the header by searching for the "Status" cell, so this
    offset is safe.
    """
    if name in wb.sheetnames:
        del wb[name]
    ws = wb.create_sheet(name)

    ws["A1"] = spec["title"]
    ws["A1"].font = TITLE_FONT
    ws["A1"].alignment = Alignment(vertical="center")
    ws["A2"] = spec["subtitle"]
    ws["A2"].font = SUB_FONT

    for offset, (key, value) in enumerate(PREAMBLE):
        ws.cell(row=4 + offset, column=1, value=key).font = SUB_FONT
        ws.cell(row=4 + offset, column=2, value=value).font = SUB_FONT

    header_row = 10
    for col, header in enumerate(spec["columns"], start=1):
        c = ws.cell(row=header_row, column=col, value=header)
        c.fill, c.font, c.alignment = HEADER_FILL, HEADER_FONT, HEADER_ALIGN

    for r, row in enumerate(spec["rows"], start=header_row + 1):
        for col, value in enumerate(row, start=1):
            c = ws.cell(row=r, column=col, value=value)
            c.alignment = BODY_ALIGN

    note_row = header_row + len(spec["rows"]) + 2
    nc = ws.cell(row=note_row, column=1, value="Note: " + NOTES)
    nc.alignment = BODY_ALIGN
    nc.font = SUB_FONT

    for col, width in enumerate(spec["widths"], start=1):
        ws.column_dimensions[ws.cell(row=header_row, column=col).column_letter].width = width
    ws.freeze_panes = f"A{header_row + 1}"


def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--dry-run", action="store_true")
    args = ap.parse_args()

    wb = openpyxl.load_workbook(WORKBOOK)
    before = list(wb.sheetnames)

    write_sheet(wb, "Form Factor", FORM_FACTOR)

    # Form Factor sits after Wire Egress: both are mechanical/installation
    # axes read against the electrical ones above them.
    order = [
        "Legend",
        "Voltage",
        "Amperage",
        "Motor",
        "Shaft Sensor",
        "Protocol",
        "Control",
        "EMI Hardening",
        "Wire Egress",
        "Form Factor",
    ]
    wb._sheets = [wb[n] for n in order if n in wb.sheetnames] + [
        wb[n] for n in wb.sheetnames if n not in order
    ]

    print(f"reference geometry: R = {REF_RADIUS_MM} mm, s = {REF_DEPTH_MM} mm")
    print(f"  exact max strip width  = {REF_WIDTH_MM:.4f} mm")
    print(f"  nominal strip width    = {NOMINAL_WIDTH_MM:.2f} mm")
    print(f"  sagitta at nominal     = {NOMINAL_SAGITTA_MM:.4f} mm")
    print(f"  facet arc at nominal   = {NOMINAL_FACET_DEG:.2f} deg")
    print(f"\nsheets before: {before}")
    print(f"sheets after:  {wb.sheetnames}")
    if args.dry_run:
        print("dry run -- nothing written")
        return 0

    stamp = datetime.now(timezone.utc).strftime("%Y%m%dT%H%M%SZ")
    backup = WORKBOOK.with_suffix(f".xlsx.{stamp}.bak")
    shutil.copy2(WORKBOOK, backup)
    wb.save(WORKBOOK)
    print(f"backed up  {backup.name}")
    print(f"wrote      {WORKBOOK}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
