#!/usr/bin/env python3
"""Add the Wire Egress axis sheet to docs/decision-matrix.xlsx.

Governed by AGENTS.md. Companion to
docs/design-single-end-wire-egress-variant.md.

WHY THIS AXIS EXISTS
--------------------
Where the five 50 A power conductors leave the board is not a layout detail
that can be settled after placement. It decides:

  * which board face each terminal group sits on, and therefore whether a
    terminal sits on its own current-carrying pour or has to reach it
    through a via field (see docs/tools/conductor_sizing.py -- the
    pour-edge-to-terminal gap costs more power than all six FETs' conduction
    loss combined if it is crossed by track instead of pour);
  * whether the inner GND planes have to act as an interface shield between
    the two faces, which in turn constrains where phase vias may transition;
  * whether the pack and phase harnesses leave in one bundle, which changes
    the conducted-emissions picture the EMI Hardening axis is sized against.

Two options are carried:

  * **Opposite-end** -- the as-built arrangement on
    builds/6s/50A/CAN_485_faraday/: pack in at one board end, phases out at
    the other. No interface shield needed; the harnesses separate naturally.
  * **Same-end, opposite faces** -- the pocket-mount variant: both groups
    leave the same board end, phases on F.Cu and pack on B.Cu, with the two
    inner GND planes carrying the shield between them.

A third arrangement -- same end, both groups on the outer faces but split
onto the two long edges -- was considered and rejected on arithmetic; see
the design document Sec.7.2. It is recorded in this sheet's notes rather
than as a selectable row, because it is not buildable at this board width.

WHAT THIS SCRIPT WILL NOT DO
-----------------------------
It does not invent part numbers or fab capabilities. AGENTS.md Sec.1.3
forbids filling a cell with a plausible-sounding answer, and two cells below
genuinely have no verified answer in this repo yet:

  * No common-mode choke has been sourced or verified. The same-end row's
    BOM cell reads TBD, not a guess.
  * The board has NO STACKUP DEFINED -- the .kicad_pcb carries four copper
    layers and `(thickness 1.6)` in its `general` block, which is KiCad's
    default rather than a design decision. No dielectric height, copper
    weight, or laminate permittivity has been specified anywhere in this
    repository, so no separation figure is asserted here.

The script is idempotent: re-running replaces the sheet rather than
duplicating it, and it backs the workbook up to a timestamped file first so
the existing decision-matrix.xlsx.bak is not clobbered.

Requires: openpyxl

Usage:
    python3 docs/tools/add_wire_egress_sheet.py [--dry-run]

Then regenerate the JSON export:
    python3 docs/tools/decision_matrix_to_json.py
"""
# Authored by Claude Opus 5 (Anthropic) 2026-08-31. AI-generated; every part
# selection is carried from this repo's own verified records or explicitly
# left open. Not human-authored. Not yet human-reviewed.

import argparse
import shutil
from datetime import datetime, timezone
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

REPO = Path(__file__).resolve().parents[2]
WORKBOOK = REPO / "docs" / "decision-matrix.xlsx"

HEADER_FILL = PatternFill("solid", fgColor="FF2F5597")
HEADER_FONT = Font(bold=True, size=10, color="FFFFFFFF")
TITLE_FONT = Font(bold=True, size=16)
SUB_FONT = Font(size=10)
BODY_ALIGN = Alignment(wrap_text=True, vertical="top")
HEADER_ALIGN = Alignment(wrap_text=True, vertical="center")

TBD = "TBD -- requires primary-source verification (AGENTS.md Sec.1.3)"

# ---------------------------------------------------------- Wire Egress -----
EGRESS = {
    "title": "Wire Egress Decision Matrix",
    "subtitle": "Build axis: Wire Egress (Opposite-end / Same-end, "
    "opposite faces)",
    "widths": [24, 30, 34, 38, 30, 44, 18, 22],
    "columns": [
        "Egress Pattern",
        "Terminal Face Assignment",
        "Board Re-partition Required",
        "Inner-Plane Shield Requirement",
        "Additional BOM Component",
        "Workflow Steps",
        "REFERENCES.md Tags",
        "Status",
    ],
    "rows": [
        [
            "Opposite-end (default)",
            (
                "Pack J5A/J5B on F.Cu at one board end; phases J4A/J4B/J4C on "
                "F.Cu at the other. Both groups sit directly on their own "
                "pours."
            ),
            (
                "None. This is the as-built arrangement on "
                "builds/6s/50A/CAN_485_faraday/."
            ),
            (
                "None arising from egress. The In1.Cu GND plane and the "
                "In2.Cu GND pour above y 60.5 mm serve their normal return "
                "and shielding roles; no terminal group is stacked over "
                "another."
            ),
            "None beyond the selected EMI Hardening tier",
            (
                "No egress-specific step. Terminal placement follows the "
                "power-stage placement; verify each terminal lands on its "
                "own pour per docs/tools/conductor_sizing.py."
            ),
            "-",
            "Candidate (unverified)",
        ],
        [
            "Same-end, opposite faces",
            (
                "Both groups leave the same board end. Phases J4A/J4B/J4C "
                "stay on F.Cu on their existing pours (unchanged); pack "
                "J5A/J5B move to B.Cu directly opposite. The two inner GND "
                "planes lie between them."
            ),
            (
                "Yes. B.Cu at the terminal end currently holds the MCU U1, "
                "the secure element U2 and the probe pads J1; all three must "
                "vacate to the logic end. VM distribution must also be "
                "extended to reach a B.Cu pad at that end -- today the VM "
                "pour stops at y 60.3 mm (F.Cu) / 60.5 mm (In2.Cu)."
            ),
            (
                "MANDATORY and NOT currently met. The unnamed rule area at "
                "x 20.75-51.15, y 76.50-86.55 mm sets `copperpour "
                "not_allowed` on ALL FOUR copper layers, so In1.Cu and "
                "In2.Cu are cut away exactly where the two terminal groups "
                "would face each other. That rule area must be restricted to "
                "the outer layers, the inner GND planes must pour solid "
                "through the terminal window, and phase via transitions must "
                "be kept out of that window so no antipad perforates the "
                "shield."
            ),
            (
                "Common-mode choke on the pack input -- "
                + TBD
                + ". No choke has been sourced or verified in this repo."
            ),
            (
                "Define a stackup FIRST -- the board currently has none "
                "(4 copper layers declared, `(thickness 1.6)` is KiCad's "
                "default, no dielectric height or copper weight specified). "
                "Then: fix the rule area's layer scope; re-place U1/U2/J1; "
                "extend VM to B.Cu; stitch In1/In2 with a GND via fence "
                "around the terminal window; re-run conductor_sizing.py and "
                "isolation_envelope.py; negative-control every new DRC rule "
                "before trusting it."
            ),
            "[9], [49]",
            "Open / unresolved",
        ],
    ],
}

NOTES = (
    "Rejected third arrangement: same end, both groups on the outer faces "
    "but split onto the two long edges. Five terminals in one row need "
    "2 x 7.0 mm + 3 x 5.0 mm = 29.0 mm of pad on a 32.00 mm edge, leaving "
    "0.50 mm per gap once four inter-pad gaps and two edge margins are "
    "taken -- not dressable with 6 sq.mm (~10 AWG) conductor, its solder "
    "fillet and its strain relief. Widening the board to suit is governed "
    "by docs/solutions/architecture-patterns/"
    "isolation-geometry-sets-board-aspect.md, which prices width at 3.0x "
    "the length it saves. See docs/design-single-end-wire-egress-variant.md "
    "Sec.7.2. -- Board thickness is NOT a usable lever for face-to-face "
    "separation: with a grounded plane between the faces the direct "
    "coupling path is already terminated, whereas doubling the laminate "
    "thickness only halves a parallel-plate capacitance. Restore the plane "
    "rather than thicken the board."
)


def write_sheet(wb, name: str, spec: dict) -> None:
    """Create (or replace) one axis sheet in the shipped house style."""
    if name in wb.sheetnames:
        del wb[name]
    ws = wb.create_sheet(name)

    ws["A1"] = spec["title"]
    ws["A1"].font = TITLE_FONT
    ws["A1"].alignment = Alignment(vertical="center")
    ws["A2"] = spec["subtitle"]
    ws["A2"].font = SUB_FONT

    for col, header in enumerate(spec["columns"], start=1):
        c = ws.cell(row=4, column=col, value=header)
        c.fill, c.font, c.alignment = HEADER_FILL, HEADER_FONT, HEADER_ALIGN

    for r, row in enumerate(spec["rows"], start=5):
        for col, value in enumerate(row, start=1):
            c = ws.cell(row=r, column=col, value=value)
            c.alignment = BODY_ALIGN

    note_row = 5 + len(spec["rows"]) + 1
    nc = ws.cell(row=note_row, column=1, value="Note: " + NOTES)
    nc.alignment = BODY_ALIGN
    nc.font = SUB_FONT

    for col, width in enumerate(spec["widths"], start=1):
        ws.column_dimensions[ws.cell(row=4, column=col).column_letter].width = width
    ws.freeze_panes = "A5"


def main() -> int:
    """Add the Wire Egress sheet to the workbook."""
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--dry-run", action="store_true")
    args = ap.parse_args()

    wb = openpyxl.load_workbook(WORKBOOK)
    before = list(wb.sheetnames)

    write_sheet(wb, "Wire Egress", EGRESS)

    # Keep axis sheets in build-walk order, Legend first. Wire Egress sits
    # after EMI Hardening: it is a mechanical/installation axis, and its
    # shield requirement is read against whatever EMI tier was chosen.
    order = [
        "Legend",
        "Voltage",
        "Amperage",
        "Motor",
        "Shaft Sensor",
        "Protocol",
        "Control",
        "EMI Hardening",
        "Wire Egress",
    ]
    wb._sheets = [wb[n] for n in order if n in wb.sheetnames] + [
        wb[n] for n in wb.sheetnames if n not in order
    ]

    print(f"sheets before: {before}")
    print(f"sheets after:  {wb.sheetnames}")
    if args.dry_run:
        print("dry run -- nothing written")
        return 0

    # Timestamped backup: decision-matrix.xlsx.bak is an existing artifact
    # from an earlier session and must not be clobbered.
    stamp = datetime.now(timezone.utc).strftime("%Y%m%dT%H%M%SZ")
    backup = WORKBOOK.with_suffix(f".xlsx.{stamp}.bak")
    shutil.copy2(WORKBOOK, backup)
    wb.save(WORKBOOK)
    print(f"backed up  {backup.name}")
    print(f"wrote      {WORKBOOK}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
