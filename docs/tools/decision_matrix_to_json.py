#!/usr/bin/env python3
"""Export docs/decision-matrix.xlsx to a scriptable JSON database.

Governed by AGENTS.md. The workbook is the human-editable source of truth;
this produces `docs/decision-matrix.json` so a build instantiator can walk the
axes programmatically instead of a person reading spreadsheet cells.

    xlsx  --(this script)-->  json  --(build scripts)-->  a new builds/ folder

Re-run this after ANY edit to the workbook. The JSON carries the workbook's
mtime and a content hash so a consumer can detect that it has gone stale
rather than silently building from an old matrix.

SHAPE
-----
    {
      "schema_version": 1,
      "source": {...},                  # provenance + staleness detection
      "status_legend": {                # status string -> its Legend meaning
        "Verified (local PDF)": "...", ...
      },
      "axes": {
        "motor": {
          "sheet": "Motor",
          "title": "...",
          "subtitle": "...",
          "columns": ["Motor Type", ...],
          "rows": [ {"Motor Type": "...", ...}, ... ],
          "rows_by_key": { "brushless_bldc_pmsm": {...}, ... }
        }, ...
      }
    }

`rows_by_key` is the scripting entry point: each row is keyed by a slug of its
first column, so a caller writes

    db["axes"]["motor"]["rows_by_key"]["brushless_bldc_pmsm"]["Gate Driver Part"]

rather than indexing by row number, which would break the moment a row is
inserted.

READING THE STATUS FIELD IS NOT OPTIONAL
-----------------------------------------
Every row carries a `Status`. A consumer that ignores it will happily build a
BOM out of `Open / unresolved` rows whose part cells read "TBD -- requires
primary-source verification". `unresolved_cells()` below exists so a build
script can refuse to proceed on those, per AGENTS.md Sec.1.3.

Requires: pip install openpyxl

Usage:
    python3 docs/tools/decision_matrix_to_json.py [--check]

    --check  exit 1 if the JSON is missing or stale relative to the workbook,
             without rewriting it (for CI)
"""
# Authored by Claude Opus 5 (Anthropic) 2026-08-16. AI-generated. Not
# human-authored. The DATA it exports is the workbook's, unchanged.

import argparse
import hashlib
import json
import re
from datetime import datetime, timezone
from pathlib import Path

import openpyxl

REPO = Path(__file__).resolve().parents[2]
WORKBOOK = REPO / "docs" / "decision-matrix.xlsx"
OUTPUT = REPO / "docs" / "decision-matrix.json"

SCHEMA_VERSION = 1
# House layout is title / subtitle / blank / headers, but it is NOT universal:
# the Voltage sheet carries a per-cell assumption block in rows 4-8 and puts
# its header row at 10. Locating the header by searching for the row that
# contains "Status" is what stops that sheet being silently dropped from the
# export -- which is exactly what a fixed HEADER_ROW = 4 did on first run.
HEADER_SEARCH_LIMIT = 20

UNRESOLVED_MARKERS = ("TBD --", "TBD —", "requires primary-source verification")


def slug(text: str) -> str:
    """Stable key from a row's first cell: 'Brushless (BLDC / PMSM)' -> ..."""
    s = re.sub(r"[^a-z0-9]+", "_", str(text).lower()).strip("_")
    return re.sub(r"_+", "_", s)


def find_header_row(ws) -> int | None:
    """Row number of the header, identified by its trailing "Status" column."""
    for r in range(1, min(HEADER_SEARCH_LIMIT, ws.max_row) + 1):
        for c in range(1, ws.max_column + 1):
            if str(ws.cell(row=r, column=c).value or "").strip() == "Status":
                return r
    return None


def read_preamble(ws, header_row: int) -> dict:
    """Key/value assumptions stated above the header (Voltage's cell block)."""
    params = {}
    for r in range(3, header_row):
        k = str(ws.cell(row=r, column=1).value or "").strip()
        v = ws.cell(row=r, column=2).value
        if k and v is not None and str(v).strip():
            params[k] = v if isinstance(v, (int, float)) else str(v).strip()
    return params


def sheet_to_axis(ws) -> dict | None:
    """Convert one axis worksheet into the axis dict. None if not an axis sheet."""
    header_row = find_header_row(ws)
    if header_row is None:
        return None                      # Legend and anything non-axis

    headers = [
        str(ws.cell(row=header_row, column=c).value or "").strip()
        for c in range(1, ws.max_column + 1)
    ]
    while headers and not headers[-1]:
        headers.pop()
    if not headers:
        return None

    # A real data row carries a Status. Sheets also end with free-text
    # footnotes that occupy only column A -- the Amperage sheet's shunt-part
    # note is one. Without this test that note became a "row" whose slugged
    # key was the entire sentence.
    status_idx = headers.index("Status") if "Status" in headers else None
    rows, notes = [], []
    for r in range(header_row + 1, ws.max_row + 1):
        values = [ws.cell(row=r, column=c).value for c in range(1, len(headers) + 1)]
        if not any(v is not None and str(v).strip() for v in values):
            continue
        cells = [("" if v is None else str(v).strip()) for v in values]
        has_status = (
            status_idx is not None
            and cells[status_idx]
            and cells[status_idx] != "None"
        )
        if not has_status:
            notes.append(cells[0])
            continue
        rows.append(dict(zip(headers, cells)))

    key_col = headers[0]
    return {
        "sheet": ws.title,
        "title": str(ws["A1"].value or "").strip(),
        "subtitle": str(ws["A2"].value or "").strip(),
        "header_row": header_row,
        "parameters": read_preamble(ws, header_row),
        "key_column": key_col,
        "columns": headers,
        "notes": notes,
        "rows": rows,
        "rows_by_key": {slug(row[key_col]): row for row in rows},
    }


def parse_legend(ws) -> dict:
    """Pull the status legend out of the Legend sheet."""
    legend, capturing = {}, False
    for r in range(1, ws.max_row + 1):
        a = str(ws.cell(row=r, column=1).value or "").strip()
        b = str(ws.cell(row=r, column=2).value or "").strip()
        if a == "Status legend":
            capturing = True
            continue
        if capturing:
            if not a:
                break
            if b:
                legend[a] = b
    return legend


def unresolved_cells(db: dict) -> list[dict]:
    """Every cell a build script must not silently consume.

    Returns one entry per cell that either sits on an `Open / unresolved` row
    or carries an explicit TBD marker.
    """
    out = []
    for axis_key, axis in db["axes"].items():
        for row_key, row in axis["rows_by_key"].items():
            status = row.get("Status", "")
            for col, val in row.items():
                if col == "Status":
                    continue
                flagged = any(m in val for m in UNRESOLVED_MARKERS)
                if flagged or status == "Open / unresolved":
                    if flagged or col == axis["key_column"]:
                        out.append({
                            "axis": axis_key, "row": row_key,
                            "column": col, "status": status,
                            "reason": "explicit TBD" if flagged else status,
                        })
    return out


def build_db() -> dict:
    """Read the workbook and return the full database dict."""
    wb = openpyxl.load_workbook(WORKBOOK, data_only=True)
    raw = WORKBOOK.read_bytes()

    axes = {}
    for name in wb.sheetnames:
        axis = sheet_to_axis(wb[name])
        if axis is not None:
            axes[slug(name)] = axis

    return {
        "schema_version": SCHEMA_VERSION,
        "source": {
            "workbook": str(WORKBOOK.relative_to(REPO)),
            "workbook_sha256": hashlib.sha256(raw).hexdigest(),
            "workbook_mtime": datetime.fromtimestamp(
                WORKBOOK.stat().st_mtime, timezone.utc).isoformat(),
            "exported": datetime.now(timezone.utc).isoformat(),
            "exported_by": "docs/tools/decision_matrix_to_json.py",
            "governed_by": "AGENTS.md",
            "note": "Generated file -- edit the workbook, then re-run the "
                    "exporter. Do not hand-edit this JSON.",
        },
        "status_legend": parse_legend(wb["Legend"]) if "Legend" in wb.sheetnames else {},
        "axes": axes,
    }


def main() -> int:
    """Export the workbook, or check the existing export for staleness."""
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--check", action="store_true",
                    help="exit 1 if the JSON is missing or stale; do not write")
    args = ap.parse_args()

    db = build_db()

    if args.check:
        if not OUTPUT.is_file():
            print(f"STALE: {OUTPUT.relative_to(REPO)} does not exist")
            return 1
        old = json.loads(OUTPUT.read_text(encoding="utf-8"))
        old_hash = (old.get("source") or {}).get("workbook_sha256")
        if old_hash != db["source"]["workbook_sha256"]:
            print("STALE: workbook has changed since the JSON was exported; "
                  "re-run docs/tools/decision_matrix_to_json.py")
            return 1
        print(f"OK: {OUTPUT.relative_to(REPO)} matches the workbook")
        return 0

    OUTPUT.write_text(json.dumps(db, indent=2, ensure_ascii=False) + "\n",
                      encoding="utf-8")
    print(f"wrote {OUTPUT.relative_to(REPO)}")
    for key, axis in db["axes"].items():
        print(f"  {key:14} {len(axis['rows'])} rows x {len(axis['columns'])} cols")

    flagged = unresolved_cells(db)
    print(f"\n{len(flagged)} cells need primary-source verification before a "
          f"build can consume them:")
    seen = set()
    for f in flagged:
        tag = (f["axis"], f["row"])
        if tag in seen:
            continue
        seen.add(tag)
        print(f"  {f['axis']}.{f['row']}: {f['reason']}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
